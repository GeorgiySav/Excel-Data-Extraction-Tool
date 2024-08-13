import openpyxl as op
from openpyxl.utils.cell import column_index_from_string, get_column_letter
import json

from dataclasses import dataclass


@dataclass
class SheetInfo:
    name: str
    input_fp: str
    input_st: str
    output_fp: str
    output_st: str
    row_offset: int
    columns: list[str]
    enabled: bool
    mode: str
    add_name: bool
    column_offset: int

def parse_json(filepath: str) -> list[SheetInfo]:
    infos = []
    # read the json file
    with open('project_info.json', 'r') as project_file:
        project_json = json.load(project_file)
        # loop through every entry in the json
        for item in project_json:
            # extract the data and save it into a the sheet info array
            sheet_temp = SheetInfo(
                name=item['name'],
                input_fp=item['input_fp'],
                input_st=item['input_st'],
                output_fp=item['output_fp'],
                output_st=item['output_st'],
                row_offset=item['row_offset'],
                columns=item['columns'],
                enabled=item['enabled'],
                mode=item['mode'],
                add_name=item['add_name'],
                column_offset=item['column_offset'])
            infos.append(sheet_temp)
    return infos

def pull_data(info: SheetInfo):
    # check if you are allowed to pull the data
    if info.enabled == False: 
        return

    # create the workbook and sheet instances for the input and output excel workbooks
    input_workbook = op.load_workbook(filename=info.input_fp, read_only=True)
    input_sheet = input_workbook[info.input_st]

    output_workbook = op.load_workbook(filename=info.output_fp)
    output_sheet = output_workbook[info.output_st]
    
    # loop through every entry in the columns attribute
    for column in info.columns:
        print(column)
        # save the values from the offset value to the end of the column to a temporary list
        values = []
        for cell in input_sheet.iter_rows(min_row=info.row_offset,
                                     max_row=input_sheet.max_row,
                                     min_col=column_index_from_string(column),
                                     max_col=column_index_from_string(column)):
            values.append(cell[0].value)
        print(values)
        # save the temporary list in the output excel workbook
        # iterate from A1 until you find an empty column
        for col in range(1, output_sheet.max_column+2):
            print(output_sheet.cell(row=1, column=col).value)
            if output_sheet.cell(row=1, column=col).value == None:
                # paste the values in
                for ro in range(len(values)):
                    output_sheet.cell(row=ro+1, column=col).value = values[ro]
                break
        output_workbook.save(filename=info.output_fp)

def next_empty_column(info: SheetInfo):
    # create the workbook and sheet instances for the input and output excel workbooks
    input_workbook = op.load_workbook(filename=info.input_fp, read_only=True, data_only=True)
    input_sheet = input_workbook[info.input_st]

    output_workbook = op.load_workbook(filename=info.output_fp)
    output_sheet = output_workbook[info.output_st]

    # find the first empty column for the names
    first_free = output_sheet.max_column + 1
    print (first_free)

    # find the longest column in the input file
    column_length = 0
    for col in info.columns:
        y = input_sheet.max_row
        while y > 0 and input_sheet.cell(row=y, column=column_index_from_string(col)).value == None: 
            y -= 1
        column_length = max(column_length, y)
        if column_length == input_sheet.max_row:
            break
    column_length = max(1, column_length)
    print(column_length)

    if info.add_name:
        for i in range(1, column_length + 1):
            output_sheet.cell(row=i, column=first_free).value = info.name
        first_free += 1

    for col in info.columns:
        for i in range(column_length + 1):
            output_sheet.cell(row=i+1, column=first_free).value = input_sheet.cell(row=i+info.row_offset, column=column_index_from_string(col)).value
        first_free += 1

    output_workbook.save(filename=info.output_fp)

def next_empty_row(info: SheetInfo):
    # create the workbook and sheet instances for the input and output excel workbooks
    input_workbook = op.load_workbook(filename=info.input_fp, read_only=True, data_only=True)
    input_sheet = input_workbook[info.input_st]

    output_workbook = op.load_workbook(filename=info.output_fp)
    output_sheet = output_workbook[info.output_st]

    # find the first empty row in the output file
    num = len(info.columns) + (1 if info.add_name else 0)
    free_row = 0
    for i in range(info.column_offset, info.column_offset + num):
        y = output_sheet.max_row
        while y > 0 and output_sheet.cell(row=y, column=i).value == None: 
            y -= 1
        free_row = max(free_row, y+1)
        if free_row == output_sheet.max_row+1:
            break
    free_row = max(1, free_row)
    print(free_row)

    # find the longest column in the input file
    column_length = 0
    for col in info.columns:
        y = input_sheet.max_row
        while y > 0 and input_sheet.cell(row=y, column=column_index_from_string(col)).value == None: 
            y -= 1
        column_length = max(column_length, y)
        if column_length == input_sheet.max_row:
            break
    column_length = max(column_length, 1) - info.row_offset
    print(column_length)

    if info.add_name:
        for i in range(column_length):
            output_sheet.cell(row=free_row+i, column=info.column_offset).value = info.name
    for i, col in enumerate(info.columns):
        for y in range(column_length):
            output_sheet.cell(row=free_row+y, column=info.column_offset + i+(1 if info.add_name else 0)).value = input_sheet.cell(row=info.row_offset+y, column=column_index_from_string(col)).value
    
    output_workbook.save(filename=info.output_fp)

def run():
    infos = parse_json("project_info.json")
    print(infos)
    for info in infos:
        print(info)
        
        if info.enabled == False:
            continue

        if info.mode == "Next empty column":
            print("Next column")
            next_empty_column(info)
        elif info.mode == "Next empty row":
            print("Next row")
            next_empty_row(info)

'''
testInfo = SheetInfo(
    'test',
    'test_input.xlsx',
    'Sheet1',
    'test_output.xlsx',
    'Sheet1',
    1,
    ['A', 'B'],
    True,
    InsertionMode.NEXT_EMPTY_COLUMN,
    True,
    1
)

#next_empty_column(testInfo)
next_empty_row(testInfo)'''